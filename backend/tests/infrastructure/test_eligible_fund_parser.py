from pathlib import Path

from openpyxl import Workbook

from infrastructure.external.eligible_fund_parser import (
    detect_and_parse,
    parse_eligible_assets_csv,
    parse_growth_xlsx,
    parse_tsumitate_from_growth_xlsx,
    parse_tsumitate_xlsx,
)


def test_parse_eligible_assets_csv_should_normalize_fields(tmp_path: Path):
    csv_path = tmp_path / "sample.csv"
    csv_path.write_text(
        "ticker,fund_name,asset_type,trust_fee_pct\n"
        " abc123 ,Sample Fund,mutual_fund,0.123\n",
        encoding="utf-8",
    )

    rows = parse_eligible_assets_csv(csv_path)
    assert len(rows) == 1
    assert rows[0]["ticker"] == "ABC123"
    assert rows[0]["fund_name"] == "Sample Fund"
    assert rows[0]["asset_type"] == "mutual_fund"
    assert rows[0]["trust_fee_pct"] == 0.123


def test_parse_tsumitate_xlsx_should_extract_rows(tmp_path: Path):
    xlsx_path = tmp_path / "tsumitate.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append([46010])
    ws.append(["金融庁"])
    ws.append(["つみたて投資枠対象商品届出一覧（対象資産別）"])
    ws.append(["【指定インデックス投資信託】"])
    ws.append(
        [
            "単一指数・複数指数の区分(※1)",
            "国内型・海外型の区分(※2)",
            "指定指数の名称又は指定指数の数(※3)",
            "ファンド名称(※4)",
            "運用会社",
        ]
    )
    ws.append(
        [
            "単一指数（株式型）",
            "国内型",
            "TOPIX",
            "SBI・iシェアーズ・TOPIXインデックス・ファンド",
            "SBIアセットマネジメント",
        ]
    )
    wb.save(xlsx_path)

    rows = parse_tsumitate_xlsx(xlsx_path)
    assert len(rows) == 1
    assert rows[0]["ticker"] == "SBI・iシェアーズ・TOPIXインデックス・ファンド"
    assert rows[0]["fund_name"] == "SBI・iシェアーズ・TOPIXインデックス・ファンド"


def test_parse_tsumitate_xlsx_multi_sheet_different_layouts(tmp_path: Path):
    xlsx_path = tmp_path / "tsumitate-multi-sheet.xlsx"
    wb = Workbook()

    index_sheet = wb.active
    assert index_sheet is not None
    index_sheet.title = "指定インデックス投資信託"
    index_sheet.append([46010])
    index_sheet.append(["金融庁"])
    index_sheet.append(["つみたて投資枠対象商品届出一覧（対象資産別）"])
    index_sheet.append(["【指定インデックス投資信託】"])
    index_sheet.append(
        [
            "単一指数・複数指数の区分(※1)",
            "国内型・海外型の区分(※2)",
            "指定指数の名称又は指定指数の数(※3)",
            "ファンド名称(※4)",
            "運用会社",
        ]
    )
    index_sheet.append(
        [
            "単一指数（株式型）",
            "国内型",
            "TOPIX",
            "SBI・iシェアーズ・TOPIXインデックス・ファンド",
            "SBIアセットマネジメント",
        ]
    )

    active_sheet = wb.create_sheet("指定インデックス投資信託以外")
    active_sheet.append([46010])
    active_sheet.append(["金融庁"])
    active_sheet.append(["つみたて投資枠対象商品届出一覧（対象資産別）"])
    active_sheet.append(["【指定インデックス投資信託以外】"])
    active_sheet.append(
        [
            "国内型・海外型の区分(※1)",
            "投資の対象としていた資産の区分(※2)",
            "ファンド名称(※3)",
            "運用会社",
        ]
    )
    active_sheet.append(
        [
            "国内型",
            "株式",
            "年金積立 Jグロース",
            "レオス・キャピタルワークス㈱",
        ]
    )
    wb.save(xlsx_path)

    rows = parse_tsumitate_xlsx(xlsx_path)
    tickers = [row["ticker"] for row in rows]
    assert "SBI・iシェアーズ・TOPIXインデックス・ファンド" in tickers
    assert "年金積立 Jグロース" in tickers
    assert "レオス・キャピタルワークス㈱" not in tickers
    assert len(tickers) == len(set(tickers))


def test_parse_growth_xlsx_should_extract_rows(tmp_path: Path):
    xlsx_path = tmp_path / "growth.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["上場株式等の投資信託等"])
    ws.append(
        [
            "リスト更新日",
            "追加・変更の別",
            "上場投信・上場投資法人の別",
            "銘柄コード",
            "ファンド名称",
            "運用会社名",
            "設定日・設立日",
            "成長投資枠取扱可能日",
            "決算回数",
            "つみたて投資枠の対象・非対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "上場投信",
            "14980",
            "Ｏｎｅ ＥＴＦ ＥＳＧ",
            "アセットマネジメントOne株式会社",
            20171127,
            20240104,
            "年2回",
            "非対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "不動産投資信託",
            "13430",
            "ＮＥＸＴ ＦＵＮＤＳ 東証ＲＥＩＴ指数連動型上場投信",
            "野村アセットマネジメント株式会社",
            20080625,
            20240104,
            "年4回",
            "非対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "上場投資法人",
            "13280",
            "ＫＤＸ不動産投資法人",
            "ケネディクス不動産投資顧問株式会社",
            20060531,
            20240104,
            "年2回",
            "非対象",
        ]
    )
    wb.save(xlsx_path)

    rows = parse_growth_xlsx(xlsx_path)
    by_ticker = {row["ticker"]: row for row in rows}
    assert len(rows) == 3
    assert by_ticker["1498.T"]["asset_type"] == "etf"
    assert by_ticker["1498.T"]["is_tsumitate_target"] is False
    assert by_ticker["1343.T"]["asset_type"] == "reit"
    assert by_ticker["1343.T"]["is_tsumitate_target"] is False
    assert by_ticker["1328.T"]["asset_type"] == "reit"
    assert by_ticker["1328.T"]["is_tsumitate_target"] is False


def test_parse_tsumitate_from_growth_xlsx_should_filter_target_rows(tmp_path: Path):
    xlsx_path = tmp_path / "growth-with-tsumitate-flag.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["一般投資家向け成長投資枠対象商品リスト（非上場ファンド）"])
    ws.append(
        [
            "リスト更新日",
            "追加・変更の別",
            "投信協会ファンドコード",
            "ファンド名称",
            "運用会社名",
            "設定日",
            "償還日",
            "成長投資枠取扱可能日",
            "決算回数",
            "つみたて投資枠の対象・非対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "AY311189",
            "ａｕスマート・ベーシック（安定）",
            "ａｕアセットマネジメント株式会社",
            20180919,
            None,
            20240104,
            "年1回",
            "対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "4431207B",
            "ＢＮＰパリバ・ブラジル・ファンド（株式型）",
            "ＢＮＰパリバ・アセットマネジメント株式会社",
            20071116,
            None,
            20240104,
            "年2回",
            "非対象",
        ]
    )
    wb.save(xlsx_path)

    rows = parse_tsumitate_from_growth_xlsx(xlsx_path)
    assert len(rows) == 1
    assert rows[0]["ticker"] == "AY311189"
    assert rows[0]["fund_name"] == "ａｕスマート・ベーシック（安定）"
    assert rows[0]["is_tsumitate_target"] is True


def test_detect_and_parse_should_route_by_extension(tmp_path: Path):
    csv_path = tmp_path / "manual.csv"
    csv_path.write_text(
        "ticker,fund_name,asset_type,trust_fee_pct\n1306.T,NEXT FUNDS TOPIX,etf,\n",
        encoding="utf-8",
    )

    rows = detect_and_parse(csv_path, wrapper="nisa_growth")
    assert len(rows) == 1
    assert rows[0]["ticker"] == "1306.T"


def test_detect_and_parse_should_prefer_growth_format_for_tsumitate(tmp_path: Path):
    xlsx_path = tmp_path / "tsumitate-via-growth-format.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["一般投資家向け成長投資枠対象商品リスト（非上場ファンド）"])
    ws.append(
        [
            "リスト更新日",
            "追加・変更の別",
            "投信協会ファンドコード",
            "ファンド名称",
            "運用会社名",
            "設定日",
            "償還日",
            "成長投資枠取扱可能日",
            "決算回数",
            "つみたて投資枠の対象・非対象",
        ]
    )
    ws.append(
        [
            20230621,
            "追加",
            "89311199",
            "ＳＢＩ・Ｖ・Ｓ＆Ｐ５００インデックス・ファンド",
            "ＳＢＩアセットマネジメント株式会社",
            20190926,
            None,
            20240104,
            "年1回",
            "対象",
        ]
    )
    wb.save(xlsx_path)

    rows = detect_and_parse(xlsx_path, wrapper="nisa_tsumitate")
    assert len(rows) == 1
    assert rows[0]["ticker"] == "89311199"
