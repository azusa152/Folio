"""Parsers for NISA eligible fund source files (CSV / XLSX)."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_CSV_TICKER_KEYS = (
    "ticker",
    "ティッカー",
    "銘柄コード",
    "ファンドコード",
    "isin",
    "isinコード",
    "isin code",
)
_CSV_FUND_NAME_KEYS = ("fund_name", "ファンド名", "商品名", "銘柄名", "名称")
_CSV_ASSET_TYPE_KEYS = ("asset_type", "資産種別", "種別")
_CSV_TRUST_FEE_KEYS = (
    "trust_fee_pct",
    "信託報酬",
    "信託報酬率",
    "実質信託報酬率",
    "信託報酬（年率）",
)


def _normalize_key(value: str) -> str:
    return value.strip().lower().replace("　", " ").replace("\n", "")


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace("%", "").replace("％", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _extract_with_keys(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def _normalize_row(
    row: dict[str, Any], *, default_asset_type: str
) -> dict[str, Any] | None:
    ticker_raw = _extract_with_keys(row, _CSV_TICKER_KEYS)
    ticker = str(ticker_raw or "").strip().upper()
    if not ticker:
        return None
    fund_name_raw = _extract_with_keys(row, _CSV_FUND_NAME_KEYS)
    asset_type_raw = _extract_with_keys(row, _CSV_ASSET_TYPE_KEYS)
    trust_fee_raw = _extract_with_keys(row, _CSV_TRUST_FEE_KEYS)
    return {
        "ticker": ticker,
        "fund_name": str(fund_name_raw or "").strip(),
        "asset_type": str(asset_type_raw or default_asset_type).strip()
        or default_asset_type,
        "trust_fee_pct": _as_float(trust_fee_raw),
    }


def parse_eligible_assets_csv(
    path: str | Path,
    *,
    default_asset_type: str = "mutual_fund",
) -> list[dict[str, Any]]:
    file_path = Path(path)
    rows: list[dict[str, Any]] = []
    with file_path.open(encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for raw_row in reader:
            normalized_row = {
                _normalize_key(str(k)): v for k, v in raw_row.items() if k
            }
            row = _normalize_row(normalized_row, default_asset_type=default_asset_type)
            if row:
                rows.append(row)
    return rows


def _iter_rows(path: str | Path) -> list[list[Any]]:
    """Return all rows from all sheets flattened (used for single-sheet files)."""
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    rows: list[list[Any]] = []
    for sheet in wb.worksheets:
        rows.extend(list(values) for values in sheet.iter_rows(values_only=True))
    return rows


def _iter_sheets(path: str | Path) -> list[list[list[Any]]]:
    """Return rows per sheet as a list of row-lists (one entry per sheet)."""
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    return [
        [list(row) for row in sheet.iter_rows(values_only=True)]
        for sheet in wb.worksheets
    ]


def _find_header_index(
    rows: list[list[Any]], required_keys: tuple[str, ...]
) -> int | None:
    normalized_required = [_normalize_key(item) for item in required_keys]
    for idx, row in enumerate(rows):
        normalized = [_normalize_key(str(v)) for v in row if v not in (None, "")]
        if all(any(req in cell for cell in normalized) for req in normalized_required):
            return idx
    return None


def is_tsumitate_asset_class_xlsx(path: str | Path) -> bool:
    """Check whether XLSX appears to be the FSA asset-class formatted sheet."""
    rows = _iter_rows(path)
    return (
        _find_header_index(rows, ("単一指数・複数指数の区分", "ファンド名称"))
        is not None
    )


def _normalize_growth_listed_code(raw_code: str) -> str:
    code = raw_code.strip().upper()
    if re.fullmatch(r"\d{5}", code) and code.endswith("0"):
        return f"{code[:4]}.T"
    if re.fullmatch(r"\d{4}", code):
        return f"{code}.T"
    return code


def _parse_tsumitate_sheet(sheet_rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Parse one sheet from an FSA Tsumitate XLSX into normalized rows."""
    header_idx = _find_header_index(sheet_rows, ("ファンド名称", "運用会社"))
    if header_idx is None:
        return []
    header = [
        _normalize_key(str(v)) if v is not None else "" for v in sheet_rows[header_idx]
    ]
    try:
        name_col_idx = next(i for i, h in enumerate(header) if "ファンド名称" in h)
    except StopIteration:
        return []

    parsed: list[dict[str, Any]] = []
    for row in sheet_rows[header_idx + 1 :]:
        if len(row) <= name_col_idx:
            continue
        fund_name = str(row[name_col_idx] or "").strip()
        if not fund_name:
            continue
        parsed.append(
            {
                "ticker": fund_name,
                "fund_name": fund_name,
                "asset_type": "mutual_fund",
                "trust_fee_pct": None,
            }
        )
    return parsed


def parse_tsumitate_xlsx(path: str | Path) -> list[dict[str, Any]]:
    """Parse FSA Tsumitate XLSX into normalized eligible-asset rows.

    The FSA file contains multiple sheets (index funds, active funds, ETFs), each
    with its own header row and column layout.  We parse every sheet independently
    so the correct fund-name column is detected per sheet.

    Note: FSA source does not expose a machine ticker/fund code column.
    We store `fund_name` as ticker surrogate to keep list/picker usable.
    """
    all_sheets = _iter_sheets(path)
    seen: dict[str, dict[str, Any]] = {}
    for sheet_rows in all_sheets:
        for row in _parse_tsumitate_sheet(sheet_rows):
            seen[row["ticker"]] = row
    return list(seen.values())


def _parse_growth_sheet(sheet_rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Parse one sheet from a Growth NISA XLSX into normalized rows."""
    header_idx = _find_header_index(sheet_rows, ("ファンド名称",))
    if header_idx is None:
        return []
    header = [
        _normalize_key(str(v)) if v is not None else "" for v in sheet_rows[header_idx]
    ]

    def _index_of(keyword: str) -> int | None:
        for i, h in enumerate(header):
            if keyword in h:
                return i
        return None

    name_idx = _index_of("ファンド名称")
    code_idx = _index_of("銘柄コード")
    if code_idx is None:
        code_idx = _index_of("投信協会ファンドコード")
    type_idx = _index_of("上場投信・上場投資法人の別")
    tsumitate_target_idx = _index_of("つみたて投資枠の対象・非対象")
    if name_idx is None or code_idx is None:
        return []

    parsed: list[dict[str, Any]] = []
    for row in sheet_rows[header_idx + 1 :]:
        if len(row) <= max(name_idx, code_idx):
            continue
        fund_name = str(row[name_idx] or "").strip()
        code_raw = str(row[code_idx] or "").strip()
        if not fund_name or not code_raw:
            continue
        # Growth list includes Tsumitate-target flag; keep all for Growth wrapper.
        _ = tsumitate_target_idx
        raw_asset_type = (
            str(row[type_idx] or "").strip() if type_idx is not None else ""
        )
        asset_type = "etf" if "上場" in raw_asset_type else "mutual_fund"
        ticker = (
            _normalize_growth_listed_code(code_raw)
            if code_raw.isdigit()
            else code_raw.upper()
        )
        parsed.append(
            {
                "ticker": ticker,
                "fund_name": fund_name,
                "asset_type": asset_type,
                "trust_fee_pct": None,
            }
        )
    return parsed


def parse_growth_xlsx(path: str | Path) -> list[dict[str, Any]]:
    """Parse Growth NISA XLSX into normalized eligible-asset rows."""
    all_sheets = _iter_sheets(path)
    seen: dict[str, dict[str, Any]] = {}
    for sheet_rows in all_sheets:
        for row in _parse_growth_sheet(sheet_rows):
            seen[row["ticker"]] = row
    return list(seen.values())


def detect_and_parse(
    path: str | Path,
    *,
    wrapper: str,
) -> list[dict[str, Any]]:
    """Detect file format by extension and parse into normalized rows."""
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return parse_eligible_assets_csv(file_path)
    if suffix == ".xlsx":
        if wrapper == "nisa_tsumitate":
            return parse_tsumitate_xlsx(file_path)
        return parse_growth_xlsx(file_path)
    raise ValueError(f"Unsupported file format: {suffix}")
